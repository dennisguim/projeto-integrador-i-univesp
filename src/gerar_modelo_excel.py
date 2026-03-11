import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import os

def gerar_modelo():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Importar Funcionários"

    # Cabeçalhos
    headers = [
        "SETOR", "SIGLA", "LOTAÇÃO", "SIAPE", "NOME", 
        "JORNADA", "ESCALA", "REMOTO_INT", "REMOTO_REV",
        "CHEFIA_NOME", "CHEFIA_MATRICULA"
    ]

    # Estilo para o cabeçalho
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        
        # Ajustar largura da coluna (aproximado)
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 20

    # Exemplo de dados
    exemplo = [
        "DIVISÃO DE ENFERMAGEM", "DIENF", "CENTRO CIRÚRGICO", "1234567", "JOÃO DA SILVA",
        "40H", "DIÁRIA", "NÃO", "SEG, TER", "MARIA SOUZA", "987654"
    ]
    
    for col, value in enumerate(exemplo, 1):
        ws.cell(row=2, column=col).value = value

    # Garante que a pasta static existe
    static_dir = os.path.join(os.path.dirname(__file__), "static")
    if not os.path.exists(static_dir):
        os.makedirs(static_dir)

    # Salvar
    file_path = os.path.join(static_dir, "modelo_importacao.xlsx")
    wb.save(file_path)
    print(f"Modelo criado com sucesso em: {file_path}")
    print("Você pode preencher este arquivo e importá-lo usando o script import_excel.py")

if __name__ == "__main__":
    gerar_modelo()

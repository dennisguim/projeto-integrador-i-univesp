import os
import io
import csv
import tempfile
import openpyxl
from datetime import datetime
from flask import Flask, render_template, request, redirect, url_for, flash, make_response, session, abort
from flask_login import LoginManager, UserMixin, login_user, login_required, logout_user, current_user

# Importações para Google API
import firebase_admin
from firebase_admin import credentials, firestore
import google.oauth2.credentials
import google_auth_oauthlib.flow
from googleapiclient.discovery import build
from googleapiclient.http import MediaFileUpload

basedir = os.path.abspath(os.path.dirname(__file__))

app = Flask(__name__)

# Configurações do Google API/Firebase
# O arquivo google-credentials.json deve estar na raiz do projeto (um nível acima de src)
path_to_creds = os.path.join(os.path.abspath(os.path.join(basedir, os.pardir)), 'google-credentials.json')
if not firebase_admin._apps:
    cred = credentials.Certificate(path_to_creds)
    firebase_admin.initialize_app(cred)

db_fs = firestore.client()

# Configurações do Google API (OAuth)
SCOPES = ['https://www.googleapis.com/auth/spreadsheets', 'https://www.googleapis.com/auth/drive.file']
os.environ['OAUTHLIB_INSECURE_TRANSPORT'] = '1'
CLIENT_SECRETS_FILE = os.path.join(os.path.abspath(os.path.join(basedir, os.pardir)), 'client_secret.json')

# Configurações do app
app.config['SECRET_KEY'] = 'uma-chave-secreta-muito-segura'

# config do flask-login
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'

# --- MODELOS (Classes para Flask-Login) ----
class Usuario(UserMixin):
    def __init__(self, id, nome_usuario, senha, perfil, setor_id=None):
        self.id = str(id)
        self.nome_usuario = nome_usuario
        self.senha = senha
        self.perfil = perfil
        self.setor_id = str(setor_id) if setor_id else None

    @staticmethod
    def get(user_id):
        doc = db_fs.collection('usuarios').document(str(user_id)).get()
        if doc.exists:
            data = doc.to_dict()
            return Usuario(id=doc.id, **data)
        return None

# carregar usuário para o Flask-Login
@login_manager.user_loader
def load_user(user_id):
    return Usuario.get(user_id)

class FirestoreObj:
    def __init__(self, id, data):
        self.id = str(id)
        for key, value in data.items():
            setattr(self, key, value)

def registrar_log(acao, detalhes=""):
    usuario_nome = current_user.nome_usuario if current_user.is_authenticated else "Anônimo"
    db_fs.collection('logs').add({
        'data_hora': firestore.SERVER_TIMESTAMP,
        'usuario': usuario_nome,
        'acao': acao,
        'detalhes': detalhes
    })

# ---- ROTAS (PÁGINAS) ----
@app.route('/', methods=['GET', 'POST'])
def login():
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))
    
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')

        # Busca no Firestore
        users_ref = db_fs.collection('usuarios')
        query = users_ref.where('nome_usuario', '==', username).limit(1).get()

        if query:
            user_doc = query[0]
            data = user_doc.to_dict()
            if data['senha'] == password:
                user = Usuario(id=user_doc.id, **data)
                login_user(user)
                registrar_log("Login", f"Usuário {username} logou no sistema.")
                return redirect(url_for('dashboard'))
        
        registrar_log("Login Falhou", f"Tentativa falha de login para o usuário {username}.")
        flash('Usuário ou senha inválidos.')
    
    return render_template('login.html')

@app.route('/dashboard')
@login_required # só entra logado
def dashboard():
    # render_template busca e processa o arquivo html da pasta templates
    return render_template('dashboard.html')

@app.route('/logout')
@login_required
def logout():
    username = current_user.nome_usuario
    logout_user()
    registrar_log("Logout", f"Usuário {username} saiu do sistema.")
    return redirect(url_for('login'))

@app.route('/funcionarios', methods=['GET'])
@login_required
def listar_funcionarios():
    # Parâmetros de Filtro (Mês/Ano referência para saber quem já lançou)
    mes_ref = request.args.get('mes', 'JANEIRO')
    ano_ref = request.args.get('ano', 2026, type=int)

    # 1. Buscar todos os funcionários acessíveis ao usuário
    funcs_ref = db_fs.collection('funcionarios')
    if current_user.perfil == 'chefe' and current_user.setor_id:
        query = funcs_ref.where('setor_id', '==', str(current_user.setor_id)).get()
    else:
        query = funcs_ref.get()

    # Carregar setores para "join"
    setor_docs = db_fs.collection('setores').get()
    setores_dict = {doc.id: FirestoreObj(doc.id, doc.to_dict()) for doc in setor_docs}

    todos_funcionarios = []
    for doc in query:
        f = FirestoreObj(doc.id, doc.to_dict())
        f.setor = setores_dict.get(str(f.setor_id))
        todos_funcionarios.append(f)
        
    # Ordenar por nome
    todos_funcionarios.sort(key=lambda x: x.nome)

    # 2. Separar em duas listas: Pendentes e Concluídos
    pendentes = []
    concluidos = []

    for func in todos_funcionarios:
        # Verifica se existe frequência no Firestore
        freq_query = db_fs.collection('frequencias')\
            .where('funcionario_id', '==', str(func.id))\
            .where('mes', '==', mes_ref)\
            .where('ano', '==', ano_ref).limit(1).get()

        if freq_query:
            func.freq_registrada = FirestoreObj(freq_query[0].id, freq_query[0].to_dict())
            concluidos.append(func)
        else:
            pendentes.append(func)

    return render_template('lista_funcionarios.html', 
                           pendentes=pendentes, 
                           concluidos=concluidos,
                           mes_ref=mes_ref,
                           ano_ref=ano_ref)

@app.route('/funcionarios/frequencia/<func_id>', methods=['GET', 'POST'])
@login_required
def registrar_frequencia(func_id):
    doc = db_fs.collection('funcionarios').document(str(func_id)).get()
    if not doc.exists:
        abort(404)
    funcionario = FirestoreObj(doc.id, doc.to_dict())
    
    # Carregar setor
    setor_doc = db_fs.collection('setores').document(str(funcionario.setor_id)).get()
    if setor_doc.exists:
        funcionario.setor = FirestoreObj(setor_doc.id, setor_doc.to_dict())

    # verificacao de seguranca. chefe só lança para o seu setor
    if current_user.perfil == 'chefe' and str(funcionario.setor_id) != str(current_user.setor_id):
        flash('Acesso negado: Este funcionário não pertence ao seu setor.')
        return redirect(url_for('listar_funcionarios'))
    
    # Se receber via GET (clique no botão da lista), preenchemos o formulário
    mes_selecionado = request.args.get('mes')
    ano_selecionado = request.args.get('ano')
    
    if request.method == 'POST':
        mes = request.form.get('mes')
        ano = int(request.form.get('ano'))
        freq_int = request.form.get('frequencia_integral')
        obs = request.form.get('observacoes')

        # Verifica se já existe (Evitar duplicidade manual)
        freq_query = db_fs.collection('frequencias')\
            .where('funcionario_id', '==', str(funcionario.id))\
            .where('mes', '==', mes)\
            .where('ano', '==', ano).limit(1).get()

        if freq_query:
             # Atualiza
             db_fs.collection('frequencias').document(freq_query[0].id).update({
                 'frequencia_integral': freq_int,
                 'observacoes': obs
             })
             registrar_log("Atualizar Frequência", f"Frequência de {funcionario.nome} ({mes}/{ano}) atualizada.")
             flash(f'Frequência de {funcionario.nome} ({mes}/{ano}) atualizada!')
        else:
            # cria registro de frequencia
            db_fs.collection('frequencias').add({
                'mes': mes,
                'ano': ano,
                'frequencia_integral': freq_int,
                'observacoes': obs,
                'funcionario_id': str(funcionario.id)
            })
            registrar_log("Lançar Frequência", f"Frequência de {funcionario.nome} ({mes}/{ano}) registrada.")
            flash(f'Frequência de {funcionario.nome} ({mes}/{ano}) registrada!')
            
        # Retorna para a lista mantendo o filtro de mês/ano
        return redirect(url_for('listar_funcionarios', mes=mes, ano=ano))
    
    return render_template('registrar_frequencia.html', 
                           funcionario=funcionario,
                           mes_padrao=mes_selecionado,
                           ano_padrao=ano_selecionado)

@app.route('/funcionarios/novo', methods=['GET', 'POST'])
@login_required
def novo_funcionario():
    # Apenas o gestor pode cadastrar
    if current_user.perfil != 'gestor':
        flash('Acesso negado: Apenas gestores podem cadastrar funcionários.')
        return redirect(url_for('listar_funcionarios'))
    
    if request.method == 'POST':
        # Captura dados do formulário
        nome = request.form.get('nome').upper()
        siape = request.form.get('siape')
        setor_id = request.form.get('setor_id')
        jornada = request.form.get('jornada')
        escala = request.form.get('escala')
        remoto_integral = request.form.get('remoto_integral')

        # salva no banco
        db_fs.collection('funcionarios').add({
            'nome': nome,
            'siape': siape,
            'lotacao': request.form.get('lotacao'),
            'setor_id': str(setor_id),
            'jornada': jornada,
            'escala': escala,
            'trabalho_remoto_integral': remoto_integral,
            'dias_remoto_revezamento': "NÃO" # padrão inicial
        })
        registrar_log("Cadastrar Funcionário", f"Funcionário {nome} (SIAPE {siape}) cadastrado.")

        flash(f'Funcionário {nome} cadastrado com sucesso!')
        return redirect(url_for('listar_funcionarios'))
    
    # Sefor GET, mostra o formulário
    query = db_fs.collection('setores').get()
    setores = [FirestoreObj(doc.id, doc.to_dict()) for doc in query]
    return render_template('form_funcionario.html', setores=setores)

@app.route('/funcionarios/editar/<id>', methods=['GET', 'POST'])
@login_required
def editar_funcionario(id):
    if current_user.perfil != 'gestor':
        flash('Acesso negado.')
        return redirect(url_for('listar_funcionarios'))
    
    doc = db_fs.collection('funcionarios').document(str(id)).get()
    if not doc.exists:
        abort(404)
    func = FirestoreObj(doc.id, doc.to_dict())
    
    if request.method == 'POST':
        nome = request.form.get('nome').upper()
        siape = request.form.get('siape')
        
        db_fs.collection('funcionarios').document(str(id)).update({
            'nome': nome,
            'siape': siape,
            'lotacao': request.form.get('lotacao'),
            'setor_id': str(request.form.get('setor_id')),
            'jornada': request.form.get('jornada'),
            'escala': request.form.get('escala'),
            'trabalho_remoto_integral': request.form.get('remoto_integral')
        })
        
        registrar_log("Editar Funcionário", f"Dados do funcionário {nome} (SIAPE {siape}) atualizados.")
        flash(f'Dados de {nome} atualizados!')
        return redirect(url_for('listar_funcionarios'))

    query = db_fs.collection('setores').get()
    setores = [FirestoreObj(doc.id, doc.to_dict()) for doc in query]
    return render_template('form_funcionario.html', setores=setores, funcionario=func)

@app.route('/funcionarios/excluir/<id>')
@login_required
def excluir_funcionario(id):
    if current_user.perfil != 'gestor':
        flash('Acesso negado.')
        return redirect(url_for('listar_funcionarios'))
    
    doc = db_fs.collection('funcionarios').document(str(id)).get()
    if not doc.exists:
        abort(404)
    data = doc.to_dict()
    nome = data['nome']
    siape = data['siape']
    
    # Remove as frequências vinculadas antes de excluir o funcionário
    freqs = db_fs.collection('frequencias').where('funcionario_id', '==', str(id)).get()
    for f_doc in freqs:
        f_doc.reference.delete()
    
    db_fs.collection('funcionarios').document(str(id)).delete()
    registrar_log("Excluir Funcionário", f"Funcionário {nome} (SIAPE {siape}) removido.")
    
    flash(f'Funcionário {nome} removido do sistema.')
    return redirect(url_for('listar_funcionarios'))

@app.route('/relatorio', methods=['GET'])
@login_required
def relatorio_geral():
    # Apenas gestor pode ver relatório geral
    if current_user.perfil != 'gestor':
        flash('Acesso negado.')
        return redirect(url_for('dashboard'))

    # Valores padrão para filtros
    mes_filtro = request.args.get('mes', 'JANEIRO')
    ano_filtro = request.args.get('ano', 2026, type=int)
    setor_id = request.args.get('setor_id')
    nome_busca = request.args.get('nome', '').strip().upper()
    freq_integral_filtro = request.args.get('freq_integral', '')

    # Firestore query
    freq_ref = db_fs.collection('frequencias')\
        .where('mes', '==', mes_filtro)\
        .where('ano', '==', ano_filtro)
    
    freq_docs = freq_ref.get()
    
    resultados = []
    # Carregar todos os funcionários e setores para "join" em memória
    func_docs = db_fs.collection('funcionarios').get()
    funcs_dict = {doc.id: FirestoreObj(doc.id, doc.to_dict()) for doc in func_docs}
    
    setor_docs = db_fs.collection('setores').get()
    setores_dict = {doc.id: FirestoreObj(doc.id, doc.to_dict()) for doc in setor_docs}
    
    for f_doc in freq_docs:
        freq = FirestoreObj(f_doc.id, f_doc.to_dict())
        func = funcs_dict.get(str(freq.funcionario_id))
        
        if not func: continue
        
        # Aplicar filtros
        if setor_id and str(func.setor_id) != str(setor_id): continue
        if nome_busca and (nome_busca not in func.nome.upper() and nome_busca not in func.siape): continue
        if freq_integral_filtro and freq.frequencia_integral != freq_integral_filtro: continue
        
        # Vincular objetos para o template
        func.setor = setores_dict.get(str(func.setor_id))
        freq.funcionario = func
        resultados.append(freq)

    setores = sorted(setores_dict.values(), key=lambda x: x.nome)

    filtros_atuais = {
        'mes': mes_filtro,
        'ano': ano_filtro,
        'setor_id': setor_id,
        'nome': request.args.get('nome', ''),
        'freq_integral': freq_integral_filtro
    }

    return render_template('relatorio.html', 
                           registros=resultados, 
                           setores=setores,
                           filtros=filtros_atuais)

@app.route('/relatorio/exportar')
@login_required
def exportar_relatorio():
    if current_user.perfil != 'gestor':
        return redirect(url_for('dashboard'))

    # 1. Recuperar dados filtrados
    mes_filtro = request.args.get('mes')
    ano_filtro = request.args.get('ano', type=int)
    setor_id = request.args.get('setor_id')
    nome_busca = request.args.get('nome', '').strip().upper()
    freq_integral_filtro = request.args.get('freq_integral', '')

    # Lógica idêntica ao relatorio_geral para buscar dados
    freq_docs = db_fs.collection('frequencias')\
        .where('mes', '==', mes_filtro)\
        .where('ano', '==', ano_filtro).get()
    
    func_docs = db_fs.collection('funcionarios').get()
    funcs_dict = {doc.id: FirestoreObj(doc.id, doc.to_dict()) for doc in func_docs}
    
    setor_docs = db_fs.collection('setores').get()
    setores_dict = {doc.id: FirestoreObj(doc.id, doc.to_dict()) for doc in setor_docs}
    
    resultados = []
    for f_doc in freq_docs:
        freq = FirestoreObj(f_doc.id, f_doc.to_dict())
        func = funcs_dict.get(str(freq.funcionario_id))
        if not func: continue
        if setor_id and str(func.setor_id) != str(setor_id): continue
        if nome_busca and (nome_busca not in func.nome.upper() and nome_busca not in func.siape): continue
        if freq_integral_filtro and freq.frequencia_integral != freq_integral_filtro: continue
        
        func.setor = setores_dict.get(str(func.setor_id))
        freq.funcionario = func
        resultados.append(freq)

    # 2. Manipular Excel com Modelo Local
    caminho_modelo = os.path.join(app.static_folder, 'modelo_frequencia.xlsx')
    wb = openpyxl.load_workbook(caminho_modelo)
    ws = wb.active

    # Ajuste de linhas se necessário
    num_registros = len(resultados)
    if num_registros > 188:
        ws.insert_rows(193, num_registros - 188)

    # Preencher dados
    for i, freq in enumerate(resultados):
        row = 5 + i
        ws.cell(row=row, column=2).value = freq.funcionario.setor.nome if freq.funcionario.setor else ""
        ws.cell(row=row, column=3).value = freq.funcionario.setor.sigla if freq.funcionario.setor else ""
        ws.cell(row=row, column=4).value = freq.funcionario.lotacao
        ws.cell(row=row, column=5).value = freq.funcionario.siape
        ws.cell(row=row, column=6).value = freq.funcionario.nome
        ws.cell(row=row, column=7).value = freq.funcionario.jornada
        ws.cell(row=row, column=8).value = freq.funcionario.escala
        ws.cell(row=row, column=9).value = freq.funcionario.trabalho_remoto_integral
        ws.cell(row=row, column=10).value = freq.funcionario.dias_remoto_revezamento
        ws.cell(row=row, column=11).value = freq.frequencia_integral
        ws.cell(row=row, column=12).value = freq.observacoes

    # 3. Retornar Excel
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    response = make_response(output.read())
    response.headers["Content-Disposition"] = f"attachment; filename=frequencia_{mes_filtro}_{ano_filtro}.xlsx"
    response.headers["Content-type"] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    return response

@app.route('/setores/excluir/<id>')
@login_required
def excluir_setor(id):
    if current_user.perfil != 'gestor': return redirect(url_for('dashboard'))

    doc = db_fs.collection('setores').document(str(id)).get()
    if not doc.exists: abort(404)
    nome = doc.to_dict()['nome']

    # Verifica se tem funcionários
    funcs = db_fs.collection('funcionarios').where('setor_id', '==', str(id)).limit(1).get()
    if funcs:
        flash('Erro: Não é possível excluir um setor que possui funcionários vinculados.')
    else:
        db_fs.collection('setores').document(str(id)).delete()
        registrar_log("Excluir Setor", f"Setor {nome} removido.")
        flash(f'Setor {nome} excluído.')

    return redirect(url_for('listar_setores'))

@app.route('/setores')
@login_required
def listar_setores():
    if current_user.perfil != 'gestor': return redirect(url_for('dashboard'))
    query = db_fs.collection('setores').get()
    lista = [FirestoreObj(doc.id, doc.to_dict()) for doc in query]
    lista.sort(key=lambda x: x.nome)
    return render_template('lista_setores.html', setores=lista)

@app.route('/setores/novo', methods=['GET', 'POST'])
@login_required
def novo_setor():
    if current_user.perfil != 'gestor': return redirect(url_for('dashboard'))

    if request.method == 'POST':
        nome = request.form.get('nome')
        sigla = request.form.get('sigla')
        db_fs.collection('setores').add({
            'nome': nome,
            'sigla': sigla,
            'chefia_nome': request.form.get('chefia_nome'),
            'chefia_matricula': request.form.get('chefia_matricula')
        })
        registrar_log("Cadastrar Setor", f"Setor {nome} ({sigla}) cadastrado.")
        flash('Setor criado!')
        return redirect(url_for('listar_setores'))

    return render_template('form_setor.html')

@app.route('/setores/editar/<id>', methods=['GET', 'POST'])
@login_required
def editar_setor(id):
    if current_user.perfil != 'gestor': return redirect(url_for('dashboard'))

    doc = db_fs.collection('setores').document(str(id)).get()
    if not doc.exists: abort(404)
    setor = FirestoreObj(doc.id, doc.to_dict())

    if request.method == 'POST':
        nome = request.form.get('nome')
        sigla = request.form.get('sigla')
        db_fs.collection('setores').document(str(id)).update({
            'nome': nome,
            'sigla': sigla,
            'chefia_nome': request.form.get('chefia_nome'),
            'chefia_matricula': request.form.get('chefia_matricula')
        })

        registrar_log("Editar Setor", f"Setor {nome} ({sigla}) atualizado.")
        flash(f'Setor {nome} atualizado!')
        return redirect(url_for('listar_setores'))

    return render_template('form_setor.html', setor=setor)

@app.route('/usuarios/novo', methods=['GET', 'POST'])
@login_required
def novo_usuario():
    if current_user.perfil != 'gestor': return redirect(url_for('dashboard'))

    if request.method == 'POST':
        username = request.form.get('username')
        db_fs.collection('usuarios').add({
            'nome_usuario': username,
            'senha': request.form.get('password'),
            'perfil': request.form.get('perfil'),
            'setor_id': str(request.form.get('setor_id')) if request.form.get('setor_id') else None
        })
        registrar_log("Cadastrar Usuário", f"Usuário {username} criado.")
        flash('Usuário criado com sucesso!')
        return redirect(url_for('listar_usuarios'))

    query = db_fs.collection('setores').get()
    setores = sorted([FirestoreObj(doc.id, doc.to_dict()) for doc in query], key=lambda x: x.nome)
    return render_template('form_usuario.html', setores=setores)

@app.route('/usuarios/editar/<id>', methods=['GET', 'POST'])
@login_required
def editar_usuario(id):
    if current_user.perfil != 'gestor':
        flash('Acesso negado.')
        return redirect(url_for('dashboard'))

    doc = db_fs.collection('usuarios').document(str(id)).get()
    if not doc.exists: abort(404)
    usuario = FirestoreObj(doc.id, doc.to_dict())

    if request.method == 'POST':
        novo_username = request.form.get('username')
        novo_perfil = request.form.get('perfil')
        novo_setor_id = request.form.get('setor_id')

        # Verificar duplicidade se mudar username
        if novo_username != usuario.nome_usuario:
            exist = db_fs.collection('usuarios').where('nome_usuario', '==', novo_username).limit(1).get()
            if exist:
                flash('Erro: Este nome de usuário já está em uso.')
                return redirect(url_for('editar_usuario', id=id))

        db_fs.collection('usuarios').document(str(id)).update({
            'nome_usuario': novo_username,
            'perfil': novo_perfil,
            'setor_id': str(novo_setor_id) if novo_setor_id else None
        })

        registrar_log("Editar Usuário", f"Dados do usuário {novo_username} atualizados.")
        flash(f'Dados de {novo_username} atualizados com sucesso!')
        return redirect(url_for('listar_usuarios'))

    query = db_fs.collection('setores').get()
    setores = sorted([FirestoreObj(doc.id, doc.to_dict()) for doc in query], key=lambda x: x.nome)
    return render_template('form_usuario.html', setores=setores, usuario=usuario)

@app.route('/usuarios')
@login_required
def listar_usuarios():
    if current_user.perfil != 'gestor':
        flash('Acesso negado.')
        return redirect(url_for('dashboard'))

    query = db_fs.collection('usuarios').get()
    lista = [FirestoreObj(doc.id, doc.to_dict()) for doc in query]
    lista.sort(key=lambda x: x.nome_usuario)
    return render_template('lista_usuarios.html', usuarios=lista)

@app.route('/usuarios/excluir/<id>')
@login_required
def excluir_usuario(id):
    if current_user.perfil != 'gestor':
        flash('Acesso negado.')
        return redirect(url_for('dashboard'))

    if str(id) == str(current_user.id):
        flash('Erro: Você não pode excluir seu próprio usuário.')
        return redirect(url_for('listar_usuarios'))

    doc = db_fs.collection('usuarios').document(str(id)).get()
    if not doc.exists: abort(404)
    nome = doc.to_dict()['nome_usuario']
    db_fs.collection('usuarios').document(str(id)).delete()
    registrar_log("Excluir Usuário", f"Usuário {nome} removido.")

    flash(f'Usuário {nome} excluído com sucesso.')
    return redirect(url_for('listar_usuarios'))

@app.route('/usuarios/alterar_senha/<id>', methods=['GET', 'POST'])
@login_required
def admin_alterar_senha(id):
    if current_user.perfil != 'gestor':
        flash('Acesso negado.')
        return redirect(url_for('dashboard'))

    doc = db_fs.collection('usuarios').document(str(id)).get()
    if not doc.exists: abort(404)
    user = FirestoreObj(doc.id, doc.to_dict())

    if request.method == 'POST':
        nova_senha = request.form.get('nova_senha')
        if nova_senha:
            db_fs.collection('usuarios').document(str(id)).update({'senha': nova_senha})
            registrar_log("Alterar Senha (Admin)", f"Senha do usuário {user.nome_usuario} alterada.")
            flash(f'Senha de {user.nome_usuario} alterada com sucesso!')
            return redirect(url_for('listar_usuarios'))
        else:
            flash('A senha não pode ser vazia.')

    return render_template('alterar_senha.html', usuario=user)

@app.route('/minha_conta/alterar_senha', methods=['GET', 'POST'])
@login_required
def minha_senha():
    if request.method == 'POST':
        senha_atual = request.form.get('senha_atual')
        nova_senha = request.form.get('nova_senha')
        confirmar = request.form.get('confirmar_senha')

        if current_user.senha != senha_atual:
            flash('Senha atual incorreta.')
        elif nova_senha != confirmar:
            flash('A nova senha e a confirmação não coincidem.')
        elif not nova_senha:
             flash('A nova senha não pode ser vazia.')
        else:
            db_fs.collection('usuarios').document(str(current_user.id)).update({'senha': nova_senha})
            registrar_log("Alterar Senha", "O próprio usuário alterou sua senha.")
            flash('Sua senha foi alterada com sucesso!')
            return redirect(url_for('dashboard'))

    return render_template('minha_senha.html')

@app.route('/status_frequencia')
@login_required
def status_frequencia():
    if current_user.perfil != 'gestor':
        flash('Acesso negado.')
        return redirect(url_for('dashboard'))

    mes_ref = request.args.get('mes', 'JANEIRO')
    ano_ref = request.args.get('ano', 2026, type=int)

    setor_docs = db_fs.collection('setores').get()
    func_docs = db_fs.collection('funcionarios').get()
    freq_docs = db_fs.collection('frequencias').where('mes', '==', mes_ref).where('ano', '==', ano_ref).get()

    # Mapear funcionários por setor
    funcs_por_setor = {}
    for f in func_docs:
        s_id = str(f.to_dict().get('setor_id'))
        if s_id not in funcs_por_setor: funcs_por_setor[s_id] = []
        funcs_por_setor[s_id].append(f.id)

    # Mapear frequências por funcionário
    freqs_lancadas = {str(f.to_dict().get('funcionario_id')) for f in freq_docs}

    status_setores = []
    for s_doc in setor_docs:
        s = FirestoreObj(s_doc.id, s_doc.to_dict())
        ids_funcs = funcs_por_setor.get(str(s.id), [])
        total_funcionarios = len(ids_funcs)

        if total_funcionarios == 0:
            status_setores.append({'setor': s, 'total': 0, 'lancados': 0, 'pendentes': 0, 'concluido': True})
            continue

        lancados = sum(1 for f_id in ids_funcs if str(f_id) in freqs_lancadas)
        status_setores.append({
            'setor': s,
            'total': total_funcionarios,
            'lancados': lancados,
            'pendentes': total_funcionarios - lancados,
            'concluido': (lancados >= total_funcionarios)
        })

    status_setores.sort(key=lambda x: x['setor'].nome)
    return render_template('status_frequencia.html', status=status_setores, mes_ref=mes_ref, ano_ref=ano_ref)

@app.route('/logs')
@login_required
def listar_logs():
    if current_user.perfil != 'gestor':
        flash('Acesso negado.')
        return redirect(url_for('dashboard'))

    # Logs ordenados por data decrescente
    logs_query = db_fs.collection('logs').order_by('data_hora', direction=firestore.Query.DESCENDING).limit(500).get()
    logs = [FirestoreObj(doc.id, doc.to_dict()) for doc in logs_query]
    return render_template('lista_logs.html', logs=logs)

@app.route('/google/login')
@login_required
def google_login():
    if not os.path.exists(CLIENT_SECRETS_FILE):
        flash("Arquivo 'client_secret.json' não encontrado. Verifique o guia 'google-planilhas.md'.")
        return redirect(url_for('relatorio_geral'))

    flow = google_auth_oauthlib.flow.Flow.from_client_secrets_file(CLIENT_SECRETS_FILE, scopes=SCOPES)
    flow.redirect_uri = url_for('google_callback', _external=True)
    authorization_url, state = flow.authorization_url(access_type='offline', include_granted_scopes='true')
    session['state'] = state
    session['google_filters'] = request.args.to_dict()
    return redirect(authorization_url)

@app.route('/google/callback')
def google_callback():
    state = session.get('state')
    if not state: return redirect(url_for('relatorio_geral'))

    flow = google_auth_oauthlib.flow.Flow.from_client_secrets_file(CLIENT_SECRETS_FILE, scopes=SCOPES, state=state)
    flow.redirect_uri = url_for('google_callback', _external=True)
    flow.fetch_token(authorization_response=request.url)

    credentials = flow.credentials
    session['google_credentials'] = {
        'token': credentials.token,
        'refresh_token': credentials.refresh_token,
        'token_uri': credentials.token_uri,
        'client_id': credentials.client_id,
        'client_secret': credentials.client_secret,
        'scopes': credentials.scopes
    }
    return redirect(url_for('exportar_google_sheets', **session.get('google_filters', {})))

@app.route('/relatorio/google-sheets')
@login_required
def exportar_google_sheets():
    if 'google_credentials' not in session:
        return redirect(url_for('google_login', **request.args))

    # Reutiliza a lógica de busca do relatório geral (simplificada aqui)
    mes_filtro = request.args.get('mes')
    ano_filtro = request.args.get('ano', type=int)
    setor_id = request.args.get('setor_id')
    nome_busca = request.args.get('nome', '').strip().upper()
    freq_integral_filtro = request.args.get('freq_integral', '')

    freq_docs = db_fs.collection('frequencias').where('mes', '==', mes_filtro).where('ano', '==', ano_filtro).get()
    func_docs = db_fs.collection('funcionarios').get()
    funcs_dict = {doc.id: FirestoreObj(doc.id, doc.to_dict()) for doc in func_docs}
    setor_docs = db_fs.collection('setores').get()
    setores_dict = {doc.id: FirestoreObj(doc.id, doc.to_dict()) for doc in setor_docs}

    resultados = []
    for f_doc in freq_docs:
        freq = FirestoreObj(f_doc.id, f_doc.to_dict())
        func = funcs_dict.get(str(freq.funcionario_id))
        if not func: continue
        if setor_id and str(func.setor_id) != str(setor_id): continue
        if nome_busca and (nome_busca not in func.nome.upper() and nome_busca not in func.siape): continue
        if freq_integral_filtro and freq.frequencia_integral != freq_integral_filtro: continue
        func.setor = setores_dict.get(str(func.setor_id))
        freq.funcionario = func
        resultados.append(freq)

    if not resultados:
        flash("Nenhum registro encontrado para exportar.")
        return redirect(url_for('relatorio_geral', **request.args))

    # Manipular Excel e Upload (Usa o service account para o Drive se preferir, ou OAuth)
    caminho_modelo = os.path.join(app.static_folder, 'modelo_frequencia.xlsx')
    wb = openpyxl.load_workbook(caminho_modelo)
    ws = wb.active
    # ... (preenchimento idêntico ao exportar_relatorio) ...
    for i, freq in enumerate(resultados):
        row = 5 + i
        ws.cell(row=row, column=2).value = freq.funcionario.setor.nome if freq.funcionario.setor else ""
        ws.cell(row=row, column=3).value = freq.funcionario.setor.sigla if freq.funcionario.setor else ""
        ws.cell(row=row, column=4).value = freq.funcionario.lotacao
        ws.cell(row=row, column=5).value = freq.funcionario.siape
        ws.cell(row=row, column=6).value = freq.funcionario.nome
        ws.cell(row=row, column=7).value = freq.funcionario.jornada
        ws.cell(row=row, column=8).value = freq.funcionario.escala
        ws.cell(row=row, column=9).value = freq.funcionario.trabalho_remoto_integral
        ws.cell(row=row, column=10).value = freq.funcionario.dias_remoto_revezamento
        ws.cell(row=row, column=11).value = freq.frequencia_integral
        ws.cell(row=row, column=12).value = freq.observacoes

    with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
        temp_path = tmp.name
        wb.save(temp_path)

    try:
        creds = google.oauth2.credentials.Credentials(**session['google_credentials'])
        drive_service = build('drive', 'v3', credentials=creds)
        file_metadata = {'name': f'Frequência {mes_filtro}-{ano_filtro}', 'mimeType': 'application/vnd.google-apps.spreadsheet'}
        media = MediaFileUpload(temp_path, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', resumable=True)
        file = drive_service.files().create(body=file_metadata, media_body=media, fields='id').execute()
        flash(f'Sucesso! Planilha criada no seu Google Drive!')
    except Exception as e:
        flash(f'Erro ao salvar no Google Drive: {str(e)}')
    finally:
        if os.path.exists(temp_path): os.remove(temp_path)

    return redirect(url_for('relatorio_geral', **request.args))

if __name__ == '__main__':
    app.run(debug=True)